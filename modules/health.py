from __future__ import annotations
from pathlib import Path
from datetime import datetime,timedelta
from openpyxl import load_workbook

_workbook = None
WORKBOOK = (
    Path(__file__).resolve().parent.parent
    / "data"
    /"friday_health_core.xlsx"
)

def _load_sheets():
    global _workbook

    if _workbook is None:
        _workbook = load_workbook(WORKBOOK, data_only=True)

    return (
        _workbook["Workout Log"],
        _workbook["Body Metrics"],
        _workbook["Nutrition"],
    )

def _latest_weight(metrics_sheet):

    latest = None

    for row in metrics_sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        date, weight = row[:2]

        if date and weight is not None:
            latest = weight

    return latest

def _previous_weight(metrics_sheet):

    weights = []

    for row in metrics_sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        date, weight = row[:2]

        if date and weight is not None:
            weights.append(weight)

    if len(weights) < 2:
        return None

    return weights[-2]

def _starting_weight(metrics_sheet):

    for row in metrics_sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        date, weight = row[:2]

        if date and weight is not None:
            return weight

    return None

def _weight_trend(metrics_sheet):

    start = _starting_weight(metrics_sheet)
    current = _latest_weight(metrics_sheet)

    if start is None or current is None:
        return None

    return current - start

def _latest_workout(workout_sheet):

    latest = None

    for row in workout_sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        date, workout, duration, notes = row[:4]

        if date and workout:
            latest = (date, workout)

    return latest

def _workout_streak(workout_sheet):

    dates = []

    for row in workout_sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        date = row[0]

        if date:
            if hasattr(date, "date"):
                date = date.date()

            dates.append(date)

    if not dates:
        return 0

    dates = sorted(set(dates), reverse=True)

    today = datetime.now().date()

    if (today - dates[0]).days > 1:
        return 0

    streak = 1

    for i in range(len(dates) - 1):

        diff = (dates[i] - dates[i + 1]).days

        if diff == 1:
            streak += 1
        else:
            break

    return streak

def _today_nutrition(nutrition_sheet):

    latest = None

    for row in nutrition_sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        date, calories, protein = row[:3]

        if date:
            latest = (calories, protein)

    return latest

def generate_health_briefing():

    workout_sheet, metrics_sheet, nutrition_sheet = _load_sheets()

    briefing = []

    weight = _latest_weight(metrics_sheet)
    previous = _previous_weight(metrics_sheet)
    trend = _weight_trend(metrics_sheet) 

    if weight is not None:

        if previous is not None:

            diff = weight - previous

            if abs(diff) >= 0.1:

                direction = "gained" if diff > 0 else "lost"

                briefing.append(
                    f"Your current weight is {weight:.1f} kilograms. "
                    f"You've {direction} {abs(diff):.1f} kilograms since your last weigh-in."
                )

            else:

                briefing.append(
                    f"Your current weight is {weight:.1f} kilograms."
                )

        else:

            briefing.append(
                f"Your current weight is {weight:.1f} kilograms."
            )
        
    if trend is not None and abs(trend) >= 0.5:

        if trend > 0:
            briefing.append(
                f"You've gained {trend:.1f} kilograms since you started tracking."
            )
        else:
            briefing.append(
                f"You've lost {abs(trend):.1f} kilograms since you started tracking."
            )

    workout = _latest_workout(workout_sheet)

    if workout:

        workout_date, workout_name = workout

        today = datetime.now().date()

        if hasattr(workout_date, "date"):
            workout_date = workout_date.date()

        days = (today - workout_date).days

        if days == 0:

            briefing.append(
                f"You completed a {workout_name} workout today."
            )

        elif days == 1:

            briefing.append(
                f"Your last workout was {workout_name} yesterday."
            )

        else:

            briefing.append(
                f"It's been {days} days since your last workout."
            )

    nutrition = _today_nutrition(nutrition_sheet)

    if nutrition:

        calories, protein = nutrition

        if calories is not None:

            remaining = max(0, 3000 - calories)

            if remaining == 0:
                briefing.append(
                    "You've reached today's calorie target."
                )
            else:
                briefing.append(
                    f"You need about {remaining:.0f} more calories today."
                )

        if protein is not None:

            remaining = max(0, 120 - protein)

            if remaining == 0:
                briefing.append(
                    "You've already hit today's protein goal."
                )
            else:
                briefing.append(
                    f"You need about {remaining:.0f} more grams of protein."
                )
    
    streak = _workout_streak(workout_sheet)

    if streak >= 2:
        briefing.append(
            f"You're currently on a {streak}-day workout streak."
        )

    return briefing